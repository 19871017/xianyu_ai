"""链接导入器（link_importer）单元测试。

覆盖：通用 URL 提取、平台识别、商品 ID 去重、xlsx/csv/json/txt 多格式解析。
"""

import os
import sys
import json
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook

from engine.link_importer import (
    infer_platform,
    extract_links_from_text,
    import_links,
)


class TestInferPlatform(unittest.TestCase):
    def test_each_platform(self):
        self.assertEqual(infer_platform("https://detail.1688.com/offer/1.html"), "1688")
        self.assertEqual(infer_platform("https://item.taobao.com/item.htm?id=1"), "taobao")
        self.assertEqual(infer_platform("https://detail.tmall.com/item.htm?id=1"), "taobao")
        self.assertEqual(infer_platform("https://item.jd.com/100.html"), "jd")
        self.assertEqual(infer_platform("https://mobile.yangkeduo.com/goods.html?goods_id=1"), "pdd")

    def test_unknown_and_empty(self):
        self.assertEqual(infer_platform("https://example.com/x"), "")
        self.assertEqual(infer_platform(""), "")
        self.assertEqual(infer_platform(None), "")


class TestExtractLinksFromText(unittest.TestCase):
    def test_extract_basic(self):
        text = (
            "看看这个 https://detail.1688.com/offer/723383755110.html 不错\n"
            "还有 https://item.taobao.com/item.htm?id=666 这个"
        )
        links = extract_links_from_text(text)
        self.assertEqual(len(links), 2)
        self.assertEqual(links[0]["platform"], "1688")
        self.assertEqual(links[0]["item_id"], "723383755110")
        self.assertEqual(links[1]["platform"], "taobao")
        self.assertEqual(links[1]["item_id"], "666")

    def test_dedupe_same_item_diff_params(self):
        text = (
            "https://detail.1688.com/offer/100.html?spm=a\n"
            "https://detail.1688.com/offer/100.html?src=zhanwai&ptid=x"
        )
        links = extract_links_from_text(text)
        self.assertEqual(len(links), 1)
        self.assertEqual(links[0]["item_id"], "100")

    def test_skip_unsupported(self):
        text = "https://example.com/a https://baidu.com/b"
        self.assertEqual(extract_links_from_text(text), [])

    def test_trailing_punctuation_stripped(self):
        text = "链接：https://item.jd.com/12345.html。"
        links = extract_links_from_text(text)
        self.assertEqual(len(links), 1)
        self.assertEqual(links[0]["url"], "https://item.jd.com/12345.html")
        self.assertEqual(links[0]["item_id"], "12345")

    def test_empty(self):
        self.assertEqual(extract_links_from_text(""), [])
        self.assertEqual(extract_links_from_text(None), [])


class TestImportLinks(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_txt(self):
        p = os.path.join(self.tmp, "links.txt")
        with open(p, "w", encoding="utf-8") as f:
            f.write("https://detail.1688.com/offer/1.html\nhttps://item.jd.com/2.html\n")
        links = import_links(p)
        self.assertEqual([l["platform"] for l in links], ["1688", "jd"])

    def test_json(self):
        p = os.path.join(self.tmp, "data.json")
        data = [
            {"title": "A", "url": "https://detail.1688.com/offer/10.html"},
            {"title": "B", "detail": "https://item.taobao.com/item.htm?id=20"},
        ]
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        links = import_links(p)
        self.assertEqual(len(links), 2)
        self.assertEqual({l["platform"] for l in links}, {"1688", "taobao"})

    def test_xlsx_cell_and_hyperlink(self):
        p = os.path.join(self.tmp, "sel.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "标题"
        ws["B1"] = "链接"
        ws["A2"] = "商品1"
        ws["B2"] = "https://detail.1688.com/offer/555.html"
        # 第二行用超链接 target 承载链接
        ws["A3"] = "商品2"
        cell = ws["B3"]
        cell.value = "点击查看"
        cell.hyperlink = "https://item.jd.com/777.html"
        wb.save(p)
        links = import_links(p)
        ids = {(l["platform"], l["item_id"]) for l in links}
        self.assertIn(("1688", "555"), ids)
        self.assertIn(("jd", "777"), ids)

    def test_missing_file(self):
        with self.assertRaises(FileNotFoundError):
            import_links(os.path.join(self.tmp, "nope.txt"))


if __name__ == "__main__":
    unittest.main()
