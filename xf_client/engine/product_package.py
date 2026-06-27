"""统一商品包结构与导出工具。

目标格式参考商品搬家/铺货工具：
- 商品信息.xlsx
- 主图_1.jpg ...
- 详情页_1.jpg ...
- SKU规格图：规格名_1.jpg

各平台采集器可以只提供部分字段，本模块负责补默认值和统一字段名。
"""

from __future__ import annotations

import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - 运行时缺少依赖时给清晰错误
    Workbook = None
    load_workbook = None

from config import EXPORT_DIR
from utils.helpers import sanitize_filename


EXPORT_HEADERS = [
    "*标题",
    "货号",
    "商品属性",
    "类目",
    "品牌",
    "规格1",
    "规格2",
    "*价格",
    "库存",
    "短标题",
    "商家SKU",
    "SKU商品条形码",
    "SKU属性",
    "无理由退货",
    "支付方式限制",
    "产地",
    "发货地",
    "商品条形码",
    "商品毛重(公斤)",
    "[包装]长(mm)",
    "[包装]宽(mm)",
    "[包装]高(mm)",
]

DEFAULT_STOCK = 1000
PACKAGE_ATTR_KEY = "_full_product_package"


def _clean_text(value: Any, max_len: int = 2000) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:max_len]


def _as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "").replace("，", "")
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    if not nums:
        return 0.0
    try:
        return round(float(nums[0]), 2)
    except Exception:
        return 0.0


def _dedupe(values: list[Any]) -> list[Any]:
    seen: set[str] = set()
    out: list[Any] = []
    for value in values or []:
        if not value:
            continue
        key = json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, dict) else str(value)
        if key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def _safe_attrs(item: dict[str, Any]) -> dict[str, Any]:
    attrs = item.get("attributes") or {}
    if isinstance(attrs, str):
        try:
            attrs = json.loads(attrs)
        except Exception:
            attrs = {}
    if not isinstance(attrs, dict):
        attrs = {}
    return attrs


def _restore_package_from_attrs(item: dict[str, Any]) -> dict[str, Any]:
    attrs = _safe_attrs(item)
    saved = attrs.get(PACKAGE_ATTR_KEY)
    if isinstance(saved, dict):
        for key, value in saved.items():
            item.setdefault(key, value)
    item["attributes"] = attrs
    return item


# 常见电商品类词（闲管家分类级联搜索用，按长度降序匹配以优先命中更具体的词）。
COMMON_CATEGORY_WORDS = [
    "连衣裙", "半身裙", "百褶裙", "牛仔裤", "打底裤", "休闲裤", "运动裤", "防晒衣",
    "羽绒服", "冲锋衣", "针织衫", "卫衣", "毛衣", "马甲", "西装", "风衣", "外套",
    "衬衫", "T恤", "短袖", "长袖", "吊带", "背心", "睡衣", "内衣", "文胸", "内裤",
    "袜子", "丝袜", "围巾", "帽子", "手套", "腰带", "领带",
    "运动鞋", "板鞋", "帆布鞋", "高跟鞋", "凉鞋", "拖鞋", "靴子", "单鞋", "皮鞋",
    "双肩包", "单肩包", "斜挎包", "手提包", "钱包", "背包", "书包",
    "手机壳", "数据线", "充电器", "充电宝", "耳机", "音箱", "键盘", "鼠标", "手表",
    "水杯", "保温杯", "茶杯", "杯子", "餐具", "碗碟", "锅具", "刀具",
    "床单", "被套", "枕头", "抱枕", "毛巾", "浴巾", "地毯", "窗帘", "桌布",
    "玩具", "积木", "娃娃", "模型", "文具", "笔记本", "书包",
    "面膜", "口红", "粉底", "眼影", "香水", "洗发水", "沐浴露", "护手霜",
    "项链", "手链", "耳环", "戒指", "发箍", "发夹", "饰品",
    "套装", "裙子", "裤子", "鞋子", "包包",
    "冰箱贴", "门贴", "墙贴", "贴纸", "挂件", "挂饰", "摆件", "摆设", "装饰画",
    "对联", "春联", "福字", "灯笼", "花瓶", "相框", "钟表", "台灯", "夜灯",
    "收纳盒", "收纳箱", "纸巾盒", "垃圾桶", "衣架", "挂钩", "地垫", "脚垫",
    "鼠标垫", "杯垫", "钥匙扣", "胸针", "手机支架", "支架", "贴画", "壁画",
]
_CATEGORY_WORDS_SORTED = sorted(set(COMMON_CATEGORY_WORDS), key=len, reverse=True)


def extract_category_keyword(title: str, category: str = "") -> str:
    """从标题/类目里提取适合闲管家分类级联搜索的品类词。

    闲管家发布页的"商品分类"是级联搜索框，必须输入能命中的品类词
    （如"连衣裙"），用标题前几个字（如"2026年波"）搜不到。
    策略：
      1) 若已有 category 字段，取其最后一段（"女装/连衣裙" -> "连衣裙"）。
      2) 在标题里按品类词库匹配，命中最靠后的具体词（中文标题品类词多在末尾）。
      3) 回退到标题末尾 2-4 个中文字。
    """
    # 1) 已有类目
    cat = (category or "").strip()
    if cat:
        seg = re.split(r"[/>\\\s]+", cat)
        seg = [s for s in seg if s]
        if seg:
            return seg[-1][:8]

    t = (title or "").strip()
    if not t:
        return ""

    # 2) 品类词库匹配，取在标题中位置最靠后的命中词
    best = ""
    best_pos = -1
    for w in _CATEGORY_WORDS_SORTED:
        pos = t.rfind(w)
        if pos > best_pos:
            best_pos = pos
            best = w
    if best:
        return best

    # 3) 回退：取末尾连续中文（去掉结尾的"厂家/批发/包邮"等噪声）
    t2 = re.sub(r"(厂家|批发|包邮|现货|新款|爆款|直销|代发|一件代发)$", "", t)
    m = re.findall(r"[\u4e00-\u9fa5]+", t2)
    if m:
        return m[-1][-4:]
    return ""


def format_attributes(attrs: dict[str, Any]) -> str:
    if not attrs:
        return ""
    parts: list[str] = []
    for key, value in attrs.items():
        if str(key).startswith("_"):
            continue
        if value in (None, "", [], {}):
            continue
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        parts.append(f"{key}:{value}")
    return "；".join(parts) + ("；" if parts else "")


# 源平台域名 → 平台标识，用于卖出后回上游一键代采。
_SOURCE_PLATFORM_HOSTS = (
    ("1688.com", "1688"),
    ("taobao.com", "taobao"),
    ("tmall.com", "taobao"),
    ("jd.com", "jd"),
    ("yangkeduo.com", "pdd"),
    ("pinduoduo.com", "pdd"),
    ("goofish.com", "xianyu"),
    ("goofish.pro", "xianyu"),
)


def _infer_source_platform(url: str) -> str:
    """根据源商品链接推断平台标识，识别不出返回空串。"""
    if not url or not isinstance(url, str):
        return ""
    low = url.lower()
    for host, platform in _SOURCE_PLATFORM_HOSTS:
        if host in low:
            return platform
    return ""


def normalize_sku_list(item: dict[str, Any]) -> list[dict[str, Any]]:
    sku_list = item.get("sku_list") or item.get("skus") or []
    if isinstance(sku_list, str):
        try:
            sku_list = json.loads(sku_list)
        except Exception:
            sku_list = []
    if not isinstance(sku_list, list):
        sku_list = []

    base_price = _as_float(
        item.get("price") or item.get("new_price") or item.get("original_price")
    )
    base_stock = item.get("stock") or item.get("inventory") or DEFAULT_STOCK

    normalized: list[dict[str, Any]] = []
    for idx, sku in enumerate(sku_list):
        if not isinstance(sku, dict):
            continue
        spec1 = _clean_text(
            sku.get("spec1")
            or sku.get("规格1")
            or sku.get("name")
            or sku.get("spec")
            or sku.get("sku_name")
            or sku.get("display_name")
            or sku.get("thumbSpec")
            or "默认"
        , 120)
        spec2 = _clean_text(sku.get("spec2") or sku.get("规格2") or sku.get("sub_spec") or "", 120)
        price = _as_float(sku.get("price") or sku.get("价格") or sku.get("group_price") or base_price)
        stock = sku.get("stock") or sku.get("库存") or sku.get("quantity") or sku.get("inventory") or base_stock
        try:
            stock = int(float(str(stock).replace(",", "")))
        except Exception:
            stock = DEFAULT_STOCK

        sku_image = (
            sku.get("sku_image")
            or sku.get("sku_image_path")
            or sku.get("local_image")
            or sku.get("thumb_path")
            or ""
        )
        sku_image_url = (
            sku.get("sku_image_url")
            or sku.get("thumb_url")
            or sku.get("image_url")
            or sku.get("hd_thumb_url")
            or ""
        )

        normalized.append({
            "spec1": spec1 or "默认",
            "spec2": spec2,
            "price": price,
            "stock": stock,
            "sku_image": sku_image,
            "sku_image_url": sku_image_url,
            "merchant_sku": sku.get("merchant_sku") or sku.get("商家SKU") or "",
            "barcode": sku.get("barcode") or sku.get("SKU商品条形码") or "",
            "sku_attrs": sku.get("sku_attrs") or sku.get("SKU属性") or {},
            "source_sku_id": str(
                sku.get("source_sku_id")
                or sku.get("skuId")
                or sku.get("sku_id")
                or sku.get("specId")
                or sku.get("merchant_sku")
                or sku.get("商家SKU")
                or ""
            ),
            "source_spec": _clean_text(
                sku.get("source_spec") or sku.get("specAttrs") or "", 120
            ),
            "raw": sku.get("raw") or {},
        })

    if not normalized:
        normalized.append({
            "spec1": item.get("default_spec") or "默认",
            "spec2": "",
            "price": base_price,
            "stock": int(base_stock or DEFAULT_STOCK),
            "sku_image": "",
            "sku_image_url": "",
            "merchant_sku": "",
            "barcode": "",
            "sku_attrs": {},
            "source_sku_id": "",
            "source_spec": "",
            "raw": {},
        })

    return normalized


def ensure_full_product_package(item: dict[str, Any]) -> dict[str, Any]:
    """补齐并统一完整商品包字段。"""
    if not isinstance(item, dict):
        return item

    item = _restore_package_from_attrs(item)
    attrs = _safe_attrs(item)

    title = _clean_text(
        item.get("title")
        or item.get("ai_title")
        or item.get("original_title")
        or item.get("goods_name")
        or ""
    , 200)
    item["title"] = title
    item.setdefault("original_title", title)

    # 常用属性字段上提。
    brand = item.get("brand") or attrs.get("品牌") or attrs.get("brand") or ""
    category = item.get("category") or item.get("cat_name") or attrs.get("类目") or ""
    article_no = item.get("article_no") or attrs.get("货号") or attrs.get("商品货号") or attrs.get("款号") or ""
    origin = item.get("origin") or attrs.get("产地") or attrs.get("发货地") or ""
    ship_from = item.get("ship_from") or attrs.get("发货地") or attrs.get("发货地址") or ""

    item["brand"] = _clean_text(brand, 80)
    item["category"] = _clean_text(category, 120)
    item["category_keyword"] = item.get("category_keyword") or extract_category_keyword(
        title, item.get("category", "")
    )
    item["article_no"] = _clean_text(article_no, 80)
    item["origin"] = _clean_text(origin, 80)
    item["ship_from"] = _clean_text(ship_from, 80)

    # 来源追溯：平台 / 源商品链接 / 源商品 id（用于卖出后回上游一键代采）。
    source_url = item.get("source_url") or item.get("link") or ""
    source_item_id = str(item.get("source_item_id") or "")
    source_platform = (
        item.get("source_platform")
        or item.get("platform")
        or _infer_source_platform(source_url)
        or ""
    )
    item["source_url"] = source_url
    item["source_item_id"] = source_item_id
    item["source_platform"] = source_platform
    item["source_seller"] = item.get("source_seller") or item.get("seller") or ""

    sku_list = normalize_sku_list(item)
    item["sku_list"] = sku_list

    image_urls = item.get("image_urls") or []
    if isinstance(image_urls, str):
        try:
            image_urls = json.loads(image_urls)
        except Exception:
            image_urls = [image_urls]
    image_urls = [x for x in image_urls if isinstance(x, str) and x]

    main_image_urls = item.get("main_image_urls") or item.get("main_images_urls") or image_urls[:8]
    detail_image_urls = item.get("detail_image_urls") or item.get("detail_images_urls") or []
    if isinstance(main_image_urls, str):
        main_image_urls = [main_image_urls]
    if isinstance(detail_image_urls, str):
        detail_image_urls = [detail_image_urls]

    item["main_image_urls"] = _dedupe(main_image_urls)[:12]
    item["detail_image_urls"] = _dedupe(detail_image_urls)[:40]

    # 本地图片字段：兼容旧 local_images。
    local_images = item.get("local_images") or []
    if isinstance(local_images, str):
        try:
            local_images = json.loads(local_images)
        except Exception:
            local_images = [local_images]
    item["local_images"] = [x for x in local_images if isinstance(x, str) and x]
    item.setdefault("main_images", item["local_images"][:8])
    item.setdefault("detail_images", [])
    item.setdefault("sku_images", [])

    item.setdefault("after_sale", item.get("after_sale_service") or attrs.get("无理由退货") or attrs.get("售后") or "支持7天无理由退货(使用后不支持)")
    item.setdefault("payment_limit", item.get("payment_limit") or attrs.get("支付方式限制") or "")
    item.setdefault("barcode", item.get("barcode") or attrs.get("商品条形码") or "")
    item.setdefault("gross_weight_kg", item.get("gross_weight_kg") or attrs.get("商品毛重(公斤)") or "")
    item.setdefault("package_length_mm", item.get("package_length_mm") or attrs.get("[包装]长(mm)") or "")
    item.setdefault("package_width_mm", item.get("package_width_mm") or attrs.get("[包装]宽(mm)") or "")
    item.setdefault("package_height_mm", item.get("package_height_mm") or attrs.get("[包装]高(mm)") or "")
    item.setdefault("short_title", item.get("short_title") or "")

    # 把完整商品包存入 attributes，避免数据库老表丢字段。
    package = {
        "title": item.get("title", ""),
        "article_no": item.get("article_no", ""),
        "category": item.get("category", ""),
        "category_keyword": item.get("category_keyword", ""),
        "brand": item.get("brand", ""),
        "sku_list": item.get("sku_list", []),
        "main_images": item.get("main_images", []),
        "detail_images": item.get("detail_images", []),
        "sku_images": item.get("sku_images", []),
        "main_image_urls": item.get("main_image_urls", []),
        "detail_image_urls": item.get("detail_image_urls", []),
        "after_sale": item.get("after_sale", ""),
        "payment_limit": item.get("payment_limit", ""),
        "origin": item.get("origin", ""),
        "ship_from": item.get("ship_from", ""),
        "source": {
            "platform": item.get("source_platform", ""),
            "url": item.get("source_url", ""),
            "item_id": item.get("source_item_id", ""),
            "seller": item.get("source_seller", ""),
        },
        "barcode": item.get("barcode", ""),
        "gross_weight_kg": item.get("gross_weight_kg", ""),
        "package_length_mm": item.get("package_length_mm", ""),
        "package_width_mm": item.get("package_width_mm", ""),
        "package_height_mm": item.get("package_height_mm", ""),
    }
    attrs[PACKAGE_ATTR_KEY] = package
    item["attributes"] = attrs
    return item


def download_product_image_groups(item: dict[str, Any], download_fn, item_dir: str) -> dict[str, Any]:
    """按主图/详情图/SKU图分组下载图片。

    download_fn(url, save_dir, index) -> local_path | None
    """
    item = ensure_full_product_package(item)
    Path(item_dir).mkdir(parents=True, exist_ok=True)

    main_dir = os.path.join(item_dir, "主图")
    detail_dir = os.path.join(item_dir, "详情页")
    sku_dir = os.path.join(item_dir, "SKU图")
    for d in (main_dir, detail_dir, sku_dir):
        os.makedirs(d, exist_ok=True)

    main_images: list[str] = []
    detail_images: list[str] = []
    sku_images: list[dict[str, str]] = []

    for idx, url in enumerate(item.get("main_image_urls") or item.get("image_urls") or []):
        saved = download_fn(url, main_dir, idx)
        if saved:
            main_images.append(saved)

    for idx, url in enumerate(item.get("detail_image_urls") or []):
        saved = download_fn(url, detail_dir, idx)
        if saved:
            detail_images.append(saved)

    sku_list = normalize_sku_list(item)
    for idx, sku in enumerate(sku_list):
        url = sku.get("sku_image_url") or ""
        if not url:
            continue
        saved = download_fn(url, sku_dir, idx)
        if saved:
            sku["sku_image"] = saved
            sku_images.append({"spec1": sku.get("spec1", "默认"), "path": saved})

    # 没有分组数据时，兼容旧逻辑：把 local_images 当主图。
    if not main_images and item.get("local_images"):
        main_images = list(item.get("local_images") or [])[:8]

    item["main_images"] = _dedupe(main_images)
    item["detail_images"] = _dedupe(detail_images)
    item["sku_images"] = _dedupe(sku_images)
    item["sku_list"] = sku_list
    item["local_images"] = _dedupe(item.get("main_images", []) + item.get("detail_images", []) + [x.get("path") for x in sku_images])
    return ensure_full_product_package(item)


def _copy_image(src: str, dst: str) -> str:
    if not src or not os.path.exists(src):
        return ""
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy2(src, dst)
    return dst


def _image_ext(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    return ext if ext in (".jpg", ".jpeg", ".png", ".webp", ".gif") else ".jpg"


def _build_excel_rows(item: dict[str, Any]) -> list[list[Any]]:
    item = ensure_full_product_package(item)
    attrs = _safe_attrs(item)
    attr_text = format_attributes(attrs)
    sku_list = normalize_sku_list(item)
    rows = []
    for sku in sku_list:
        sku_attrs = sku.get("sku_attrs") or {}
        sku_attr_text = format_attributes(sku_attrs) if isinstance(sku_attrs, dict) else _clean_text(sku_attrs, 500)
        rows.append([
            item.get("title", ""),
            item.get("article_no", ""),
            attr_text,
            item.get("category", ""),
            item.get("brand", ""),
            sku.get("spec1", "默认"),
            sku.get("spec2", ""),
            sku.get("price", 0),
            sku.get("stock", DEFAULT_STOCK),
            item.get("short_title", ""),
            sku.get("merchant_sku", ""),
            sku.get("barcode", ""),
            sku_attr_text,
            item.get("after_sale", ""),
            item.get("payment_limit", ""),
            item.get("origin", ""),
            item.get("ship_from", ""),
            item.get("barcode", ""),
            item.get("gross_weight_kg", ""),
            item.get("package_length_mm", ""),
            item.get("package_width_mm", ""),
            item.get("package_height_mm", ""),
        ])
    return rows


def export_product_package(item: dict[str, Any], output_dir: str | None = None) -> str:
    """导出单个商品包，返回目录路径。"""
    if Workbook is None:
        raise RuntimeError("缺少 openpyxl，请先安装：pip install openpyxl")

    item = ensure_full_product_package(item)
    title = sanitize_filename(item.get("title") or item.get("item_id") or "商品")[:60]
    if output_dir is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(EXPORT_DIR, f"商品包_{ts}_{title}")
    os.makedirs(output_dir, exist_ok=True)

    # 复制图片到和样例一致的命名。
    for idx, src in enumerate(item.get("main_images") or item.get("local_images") or [], start=1):
        _copy_image(src, os.path.join(output_dir, f"主图_{idx}{_image_ext(src)}"))

    for idx, src in enumerate(item.get("detail_images") or [], start=1):
        _copy_image(src, os.path.join(output_dir, f"详情页_{idx}{_image_ext(src)}"))

    for idx, sku in enumerate(normalize_sku_list(item), start=1):
        src = sku.get("sku_image") or ""
        if src and os.path.exists(src):
            spec = sanitize_filename(sku.get("spec1") or f"SKU_{idx}")[:80]
            _copy_image(src, os.path.join(output_dir, f"{spec}_1{_image_ext(src)}"))

    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"
    ws.append(EXPORT_HEADERS)
    for row in _build_excel_rows(item):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 36, "B": 14, "C": 42, "D": 16, "E": 14,
        "F": 26, "G": 18, "H": 12, "I": 10, "J": 20,
        "K": 18, "L": 18, "M": 26, "N": 24, "O": 18,
        "P": 14, "Q": 14, "R": 18, "S": 14, "T": 14,
        "U": 14, "V": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    xlsx_path = os.path.join(output_dir, "商品信息.xlsx")
    wb.save(xlsx_path)
    return output_dir


def export_products_package(items: list[dict[str, Any]], output_dir: str | None = None) -> str:
    """导出多个商品包。一个商品时直接生成样例格式；多个商品时分商品子目录。"""
    items = [ensure_full_product_package(x) for x in (items or []) if isinstance(x, dict)]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if output_dir is None:
        output_dir = os.path.join(EXPORT_DIR, f"商品包_{ts}")
    os.makedirs(output_dir, exist_ok=True)

    if len(items) == 1:
        return export_product_package(items[0], output_dir=output_dir)

    for idx, item in enumerate(items, start=1):
        title = sanitize_filename(item.get("title") or item.get("item_id") or f"商品_{idx}")[:50]
        subdir = os.path.join(output_dir, f"{idx:03d}_{title}")
        export_product_package(item, output_dir=subdir)
    return output_dir


# ─────────────────────────────────────────────────────────────
#  导入：商品包目录 → 标准 item（export_product_package 的逆操作）
# ─────────────────────────────────────────────────────────────

# 表头别名 → 标准字段。兼容带 * 前缀、繁简、空格差异。
_HEADER_ALIASES = {
    "标题": "title",
    "货号": "article_no",
    "商品属性": "_attr_text",
    "类目": "category",
    "品牌": "brand",
    "规格1": "spec1",
    "规格2": "spec2",
    "价格": "price",
    "库存": "stock",
    "短标题": "short_title",
    "商家sku": "merchant_sku",
    "sku商品条形码": "barcode",
    "sku属性": "_sku_attr_text",
    "无理由退货": "after_sale",
    "支付方式限制": "payment_limit",
    "产地": "origin",
    "发货地": "ship_from",
    "商品条形码": "product_barcode",
    "商品毛重(公斤)": "gross_weight_kg",
    "[包装]长(mm)": "package_length_mm",
    "[包装]宽(mm)": "package_width_mm",
    "[包装]高(mm)": "package_height_mm",
}

_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif")


def _norm_header(name: Any) -> str:
    text = re.sub(r"\s+", "", str(name or "")).lstrip("*＊").lower()
    return text


def _parse_attr_text(text: Any) -> dict[str, str]:
    """把 ``材质:聚丙烯；风格:中式；`` 反解析为属性字典。"""
    attrs: dict[str, str] = {}
    if not text:
        return attrs
    for part in re.split(r"[；;]", str(text)):
        part = part.strip()
        if not part:
            continue
        m = re.split(r"[:：]", part, maxsplit=1)
        if len(m) == 2:
            key, val = m[0].strip(), m[1].strip()
            if key and val:
                attrs[key] = val
    return attrs


def _natural_key(name: str) -> list:
    return [int(t) if t.isdigit() else t.lower()
            for t in re.split(r"(\d+)", name)]


def _list_images(directory: str, prefixes: tuple[str, ...]) -> list[str]:
    """目录下按前缀匹配的图片，自然排序返回绝对路径。"""
    if not directory or not os.path.isdir(directory):
        return []
    out = []
    for fn in os.listdir(directory):
        low = fn.lower()
        if not low.endswith(_IMAGE_EXTS):
            continue
        if prefixes and not any(fn.startswith(p) for p in prefixes):
            continue
        out.append(fn)
    out.sort(key=_natural_key)
    return [os.path.join(directory, fn) for fn in out]


def _find_image_dir(root: str, subdir_names: tuple[str, ...]) -> str:
    """返回存在的图片子目录路径；都不存在时回退根目录。"""
    for name in subdir_names:
        cand = os.path.join(root, name)
        if os.path.isdir(cand):
            return cand
    return root


def _collect_main_images(root: str) -> list[str]:
    sub = _find_image_dir(root, ("主图", "主图片", "main", "mainimages"))
    if sub != root:
        imgs = _list_images(sub, ())
        if imgs:
            return imgs
    return _list_images(root, ("主图", "主图_"))


def _collect_detail_images(root: str) -> list[str]:
    sub = _find_image_dir(root, ("详情图", "详情页", "detail", "detailimages"))
    if sub != root:
        imgs = _list_images(sub, ())
        if imgs:
            return imgs
    return _list_images(root, ("详情页", "详情图"))


def _collect_sku_image(root: str, spec1: str) -> str:
    """按规格名匹配 SKU 图：``{规格名}_1.jpg``。优先 SKU图 子目录。"""
    if not spec1:
        return ""
    safe = sanitize_filename(spec1)
    candidates = (spec1, safe)
    search_dirs = []
    for name in ("SKU图", "sku图", "SKU", "skuimages"):
        d = os.path.join(root, name)
        if os.path.isdir(d):
            search_dirs.append(d)
    search_dirs.append(root)
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        files = [f for f in os.listdir(d) if f.lower().endswith(_IMAGE_EXTS)]
        for cand in candidates:
            for f in files:
                stem = os.path.splitext(f)[0]
                if stem == cand or stem == f"{cand}_1" or stem.startswith(f"{cand}_"):
                    return os.path.join(d, f)
    return ""


def import_product_package(package_dir: str) -> dict[str, Any]:
    """读取一个商品包目录（含 商品信息.xlsx + 图片），返回标准 item。

    是 export_product_package 的逆操作。兼容两种布局：
      1. 扁平：根目录下 主图_N / 详情页_N / {规格}_1。
      2. 子目录：主图/、详情图/、SKU图/。
    多行 = 多规格；商品级字段（标题/属性等）从首个非空行向后继承。
    """
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，请先安装：pip install openpyxl")
    if not os.path.isdir(package_dir):
        raise FileNotFoundError(f"目录不存在: {package_dir}")

    xlsx_path = None
    for fn in os.listdir(package_dir):
        if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("~$"):
            xlsx_path = os.path.join(package_dir, fn)
            if "商品信息" in fn:
                break
    if not xlsx_path:
        raise FileNotFoundError(f"目录内未找到商品信息表格: {package_dir}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("表格为空")

    header = rows[0]
    col_map: dict[int, str] = {}
    for idx, name in enumerate(header):
        field = _HEADER_ALIASES.get(_norm_header(name))
        if field:
            col_map[idx] = field

    def cell(row, field):
        for idx, f in col_map.items():
            if f == field and idx < len(row):
                return row[idx]
        return None

    product: dict[str, Any] = {}
    sku_list: list[dict[str, Any]] = []
    product_fields = (
        "title", "article_no", "category", "brand", "short_title",
        "after_sale", "payment_limit", "origin", "ship_from",
        "product_barcode", "gross_weight_kg",
        "package_length_mm", "package_width_mm", "package_height_mm",
        "_attr_text",
    )

    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        # 商品级字段：首个非空即采用（后续行通常为空，沿用首行）。
        for field in product_fields:
            val = cell(row, field)
            if val not in (None, "") and not product.get(field):
                product[field] = val

        spec1 = _clean_text(cell(row, "spec1") or "", 120)
        spec2 = _clean_text(cell(row, "spec2") or "", 120)
        price = _as_float(cell(row, "price"))
        stock_val = cell(row, "stock")
        try:
            stock = int(float(str(stock_val).replace(",", ""))) if stock_val not in (None, "") else DEFAULT_STOCK
        except Exception:
            stock = DEFAULT_STOCK
        # 价格和规格都空的行跳过（避免误把空白行当 SKU）。
        if not spec1 and not price:
            continue
        sku_list.append({
            "spec1": spec1 or "默认",
            "spec2": spec2,
            "price": price,
            "stock": stock,
            "sku_image_url": "",
            "sku_image": _collect_sku_image(package_dir, spec1),
            "merchant_sku": _clean_text(cell(row, "merchant_sku") or "", 80),
            "barcode": _clean_text(cell(row, "barcode") or "", 80),
            "sku_attrs": _parse_attr_text(cell(row, "_sku_attr_text")),
            "raw": {},
        })

    attrs = _parse_attr_text(product.get("_attr_text"))

    main_images = _collect_main_images(package_dir)
    detail_images = _collect_detail_images(package_dir)

    item: dict[str, Any] = {
        "item_id": f"import_{sanitize_filename(os.path.basename(package_dir.rstrip('/')))[:40]}",
        "platform": "import",
        "title": _clean_text(product.get("title") or "", 200),
        "article_no": _clean_text(product.get("article_no") or "", 80),
        "category": _clean_text(product.get("category") or "", 120),
        "brand": _clean_text(product.get("brand") or "", 80),
        "short_title": _clean_text(product.get("short_title") or "", 120),
        "after_sale": _clean_text(product.get("after_sale") or "", 200),
        "payment_limit": _clean_text(product.get("payment_limit") or "", 120),
        "origin": _clean_text(product.get("origin") or "", 80),
        "ship_from": _clean_text(product.get("ship_from") or "", 80),
        "barcode": _clean_text(product.get("product_barcode") or "", 80),
        "gross_weight_kg": product.get("gross_weight_kg") or "",
        "package_length_mm": product.get("package_length_mm") or "",
        "package_width_mm": product.get("package_width_mm") or "",
        "package_height_mm": product.get("package_height_mm") or "",
        "attributes": attrs,
        "sku_list": sku_list,
        "main_images": main_images,
        "local_images": main_images,
        "detail_images": detail_images,
        "image_dir": package_dir,
        "source_dir": package_dir,
    }

    prices = [s["price"] for s in sku_list if s.get("price")]
    item["price"] = min(prices) if prices else 0.0
    item["original_price"] = str(item["price"])

    return ensure_full_product_package(item)
