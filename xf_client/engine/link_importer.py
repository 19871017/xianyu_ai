"""链接导入器：从选品/导出文件中提取各平台商品链接（纯函数，便于单元测试）。

配合方案一：用 1688 官方采购助手插件(或任意来源)选品并导出文件
(Excel/CSV/JSON/TXT)，本模块从中提取受支持平台的商品链接，交给批量采集。

设计原则：
- 不绑定任何插件的具体字段结构，改用「通用 URL 提取 + 平台识别」，
  插件升级或换来源都不受影响，鲁棒性最好。
- 无浏览器依赖，全部容错；解析不到返回空列表。
"""

from __future__ import annotations

import json
import os
import re

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - 运行时缺少依赖时给清晰错误
    load_workbook = None


# 受支持平台 host → 平台标识（与 product_package._SOURCE_PLATFORM_HOSTS 对齐）
_PLATFORM_HOSTS = (
    ("1688.com", "1688"),
    ("taobao.com", "taobao"),
    ("tmall.com", "taobao"),
    ("jd.com", "jd"),
    ("yangkeduo.com", "pdd"),
    ("pinduoduo.com", "pdd"),
)

# 各平台商品 ID 提取规则（用于去重：同一商品不同参数视为一个）
_ITEM_ID_PATTERNS = (
    ("1688", re.compile(r"/offer/(\d+)")),
    ("taobao", re.compile(r"[?&]id=(\d+)")),
    ("jd", re.compile(r"/(\d+)\.html")),
    ("pdd", re.compile(r"[?&]goods_id=(\d+)")),
)

# 通用 URL 提取：到空白或常见中英文分隔/标点为止
_URL_RE = re.compile(r"https?://[^\s\"'<>）)\]】，,；;、|]+", re.I)


def infer_platform(url: str) -> str:
    """根据链接 host 推断平台标识，识别不出返回空串。"""
    if not url or not isinstance(url, str):
        return ""
    low = url.lower()
    for host, platform in _PLATFORM_HOSTS:
        if host in low:
            return platform
    return ""


def _extract_item_id(url: str, platform: str) -> str:
    for plat, pat in _ITEM_ID_PATTERNS:
        if plat == platform:
            m = pat.search(url)
            if m:
                return m.group(1)
    return ""


def _clean_url(url: str) -> str:
    """去掉 URL 尾部的标点残留。"""
    return (url or "").rstrip(".,;:，。；）)】]\"'").strip()


def extract_links_from_text(text: str) -> list[dict[str, str]]:
    """从任意文本中提取受支持平台的商品链接。

    返回 [{"url", "platform", "item_id"}]，按 (platform, item_id) 去重并保序；
    item_id 提不到时退化为按完整 url 去重。
    """
    if not text or not isinstance(text, str):
        return []
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for raw in _URL_RE.findall(text):
        url = _clean_url(raw)
        platform = infer_platform(url)
        if not platform:
            continue
        item_id = _extract_item_id(url, platform)
        key = (platform, item_id or url)
        if key in seen:
            continue
        seen.add(key)
        out.append({"url": url, "platform": platform, "item_id": item_id})
    return out


def _read_xlsx_text(path: str) -> str:
    """把 xlsx 所有单元格(含超链接 target)拼成纯文本，交给通用提取。"""
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法读取 Excel 文件")
    parts: list[str] = []
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
                hl = getattr(cell, "hyperlink", None)
                if hl is not None and getattr(hl, "target", None):
                    parts.append(str(hl.target))
    return "\n".join(parts)


def _read_json_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # 整个 JSON 转回字符串再做 URL 提取，最省心也最鲁棒。
    return json.dumps(data, ensure_ascii=False)


def import_links(path: str) -> list[dict[str, str]]:
    """主入口：按扩展名解析文件并提取商品链接。

    支持 .xlsx/.xlsm/.xls、.csv、.json、.txt 及其他文本文件。
    """
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        text = _read_xlsx_text(path)
    elif ext == ".json":
        text = _read_json_text(path)
    else:
        try:
            with open(path, "r", encoding="utf-8") as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(path, "r", encoding="gbk", errors="ignore") as f:
                text = f.read()
    return extract_links_from_text(text)
