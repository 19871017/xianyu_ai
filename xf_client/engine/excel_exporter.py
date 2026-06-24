import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.helpers import ensure_dir, sanitize_filename
from config import EXPORT_DIR


class ExcelExporter:
    """Excel结构化导出"""

    HEADERS = ["序号", "商品ID", "原始标题", "AI优化标题", "原始价格", "AI描述", "想要数", "浏览数", "收藏数", "卖家", "商品链接", "图片路径"]

    def __init__(self, save_dir: str = None):
        self.save_dir = save_dir or EXPORT_DIR

    def export(self, items: list, filename: str = None) -> str:
        if not items:
            raise ValueError("没有数据可导出")

        ensure_dir(self.save_dir)
        if not filename:
            from utils.helpers import timestamp_str
            filename = f"闲鱼数据_{timestamp_str()}.xlsx"

        filepath = os.path.join(self.save_dir, sanitize_filename(filename))

        wb = Workbook()
        ws = wb.active
        ws.title = "闲鱼商品数据"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, item in enumerate(items, 2):
            data = [
                row_idx - 1,
                item.get("item_id", ""),
                item.get("original_title", ""),
                item.get("ai_title", ""),
                item.get("original_price", ""),
                item.get("ai_description", ""),
                item.get("wants", "0"),
                item.get("views", "0"),
                item.get("collects", "0"),
                item.get("seller", ""),
                item.get("link", ""),
                ", ".join(item.get("local_images", [])),
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 35
        ws.column_dimensions['L'].width = 30

        wb.save(filepath)
        return filepath

    def preview_data(self, items: list) -> list:
        """获取预览数据"""
        return [
            {
                "序号": i + 1,
                "商品ID": item.get("item_id", ""),
                "原始标题": item.get("original_title", ""),
                "AI标题": item.get("ai_title", ""),
                "价格": item.get("original_price", ""),
                "想要": item.get("wants", "0"),
                "浏览": item.get("views", "0"),
            }
            for i, item in enumerate(items)
        ]
